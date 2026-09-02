"""Transforma prospeccao/planilha-leads.csv num workbook de trabalho:
dashboard com graficos, leads filtraveis, aba de prioridade para tocar a
operacao, lista de follow do Instagram e os avisos de verificacao.

Rodar: ~/.asdf/installs/python/3.11.9/bin/python3 scripts/gerar-xlsx.py
"""

import csv
import re
from collections import Counter
from datetime import date

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── identidade visual (mesma paleta do site, ajustada para papel branco) ─────
ESCURO = "0B1220"
CIANO = "00B8D4"
VERDE = "00A862"
AMBAR = "E8A33D"
CINZA = "8A94A6"
BRANCO = "FFFFFF"
LINHA_PAR = "F4F7FA"

FILL_ESCURO = PatternFill("solid", fgColor=ESCURO)
FILL_CIANO = PatternFill("solid", fgColor=CIANO)
FILL_QUENTE = PatternFill("solid", fgColor="D6F5E3")
FILL_MORNO = PatternFill("solid", fgColor="FDF0DC")
FILL_FRIO = PatternFill("solid", fgColor="EEF1F5")
FILL_PAR = PatternFill("solid", fgColor=LINHA_PAR)

F_TITULO = Font(name="Calibri", size=20, bold=True, color=ESCURO)
F_SUB = Font(name="Calibri", size=11, color=CINZA)
F_CAB = Font(name="Calibri", size=10, bold=True, color=BRANCO)
F_KPI_NUM = Font(name="Calibri", size=26, bold=True, color=ESCURO)
F_KPI_ROT = Font(name="Calibri", size=9, bold=True, color=CINZA)
F_SECAO = Font(name="Calibri", size=13, bold=True, color=ESCURO)
F_CORPO = Font(name="Calibri", size=10)
F_NEGRITO = Font(name="Calibri", size=10, bold=True)

BORDA_FINA = Border(*[Side(style="thin", color="DDE3EA")] * 4)
F_LINK = Font(name="Calibri", size=10, color="0563C1", underline="single")


def aplicar_link(ws, linha, coluna, destino, texto=None):
    """Deixa a celula clicavel. Sem valor na celula, nao faz nada."""
    c = ws.cell(row=linha, column=coluna)
    if texto is not None:
        c.value = texto
    if not c.value:
        return
    c.hyperlink = destino
    c.font = F_LINK

# ── precos reais configurados no painel (chave pricing_en) ──────────────────
PRECO_FULL = 1497.00
PRECO_M6_TOTAL = 299.00 * 6
PRECO_M12_TOTAL = 179.00 * 12
CARE_MENSAL = 59.00
CARE_MESES_INCLUSOS = 12


def seguidores_num(txt):
    """'1,583' -> 1583 | '11K' -> 11000 | '' -> 0"""
    if not txt:
        return 0
    t = txt.strip().replace(",", "")
    m = re.match(r"^([\d.]+)\s*([KkMm]?)$", t)
    if not m:
        return 0
    valor = float(m.group(1))
    sufixo = m.group(2).upper()
    if sufixo == "K":
        valor *= 1_000
    elif sufixo == "M":
        valor *= 1_000_000
    return int(valor)


def temperatura(pot):
    if pot >= 8:
        return "Quente"
    if pot >= 6:
        return "Morno"
    if pot >= 4:
        return "Frio"
    return "Descartar"


with open("prospeccao/planilha-leads.csv", encoding="utf-8-sig") as f:
    leads = list(csv.DictReader(f))

for l in leads:
    l["potencial"] = int(l["potencial"])
    l["nota_site"] = int(l["nota_site"])
    l["seg_num"] = seguidores_num(l["seguidores_ig"])
    l["temp"] = temperatura(l["potencial"])

leads.sort(key=lambda l: (-l["potencial"], l["nicho"]))

wb = Workbook()

# ════════════════════════════════════════════════════════════════════════════
# aba oculta com as series dos graficos
# ════════════════════════════════════════════════════════════════════════════
dados = wb.create_sheet("_dados")

por_nicho = Counter(l["nicho"] for l in leads).most_common()
por_cidade = Counter(l["cidade"] for l in leads).most_common()
por_builder = Counter(l["builder"].replace("custom/desconhecido", "sob medida") for l in leads).most_common()
por_temp = [(t, sum(1 for l in leads if l["temp"] == t)) for t in ("Quente", "Morno", "Frio", "Descartar")]

def escrever_serie(col, titulo, pares):
    dados.cell(row=1, column=col, value=titulo)
    dados.cell(row=1, column=col + 1, value="qtd")
    for i, (rotulo, qtd) in enumerate(pares, start=2):
        dados.cell(row=i, column=col, value=rotulo)
        dados.cell(row=i, column=col + 1, value=qtd)
    return len(pares)

n_nicho = escrever_serie(1, "nicho", por_nicho)
n_cidade = escrever_serie(4, "cidade", por_cidade)
n_temp = escrever_serie(7, "temperatura", por_temp)
n_builder = escrever_serie(10, "builder", por_builder)
dados.sheet_state = "hidden"

# ════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ════════════════════════════════════════════════════════════════════════════
dash = wb.active
dash.title = "Dashboard"
dash.sheet_view.showGridLines = False

dash["B2"] = "RVLand Devs · Funil de Prospecção EUA"
dash["B2"].font = F_TITULO
dash["B3"] = f"Levantado em {date.today().strftime('%d/%m/%Y')} · {len(leads)} negócios visitados, avaliados um a um pelo screenshot do site"
dash["B3"].font = F_SUB

quentes = [l for l in leads if l["potencial"] >= 8]
mornos = [l for l in leads if 6 <= l["potencial"] <= 7]
com_email = [l for l in leads if l["emails"]]
com_ig = [l for l in leads if l["instagram"]]
alcance = sum(l["seg_num"] for l in leads)

kpis = [
    ("LEADS AVALIADOS", len(leads), ESCURO),
    ("QUENTES (8-10)", len(quentes), VERDE),
    ("MORNOS (6-7)", len(mornos), AMBAR),
    ("COM E-MAIL", len(com_email), CIANO),
    ("COM INSTAGRAM", len(com_ig), CIANO),
    ("ALCANCE SOMADO", alcance, ESCURO),
]

col = 2
for rotulo, valor, cor in kpis:
    letra = get_column_letter(col)
    prox = get_column_letter(col + 1)
    dash.merge_cells(f"{letra}5:{prox}5")
    dash.merge_cells(f"{letra}6:{prox}6")
    c_rot = dash[f"{letra}5"]
    c_rot.value = rotulo
    c_rot.font = F_KPI_ROT
    c_rot.alignment = Alignment(horizontal="center")
    c_val = dash[f"{letra}6"]
    c_val.value = valor
    c_val.font = Font(name="Calibri", size=24, bold=True, color=cor)
    c_val.alignment = Alignment(horizontal="center")
    c_val.number_format = "#,##0"
    for linha in (5, 6):
        for c in (col, col + 1):
            dash.cell(row=linha, column=c).fill = FILL_FRIO
            dash.cell(row=linha, column=c).border = BORDA_FINA
    col += 3

dash.row_dimensions[5].height = 18
dash.row_dimensions[6].height = 34

# ── graficos ────────────────────────────────────────────────────────────────
dash["B9"] = "Onde estão os leads"
dash["B9"].font = F_SECAO

g_nicho = BarChart()
g_nicho.type = "bar"
g_nicho.title = "Leads por nicho"
g_nicho.height, g_nicho.width = 11, 12
g_nicho.add_data(Reference(dados, min_col=2, min_row=1, max_row=n_nicho + 1), titles_from_data=True)
g_nicho.set_categories(Reference(dados, min_col=1, min_row=2, max_row=n_nicho + 1))
g_nicho.legend = None
g_nicho.dLbls = DataLabelList()
g_nicho.dLbls.showVal = True
dash.add_chart(g_nicho, "B11")

g_temp = PieChart()
g_temp.title = "Temperatura da carteira"
g_temp.height, g_temp.width = 11, 10
g_temp.add_data(Reference(dados, min_col=8, min_row=1, max_row=n_temp + 1), titles_from_data=True)
g_temp.set_categories(Reference(dados, min_col=7, min_row=2, max_row=n_temp + 1))
g_temp.dLbls = DataLabelList()
g_temp.dLbls.showVal = True
g_temp.dLbls.showPercent = True
dash.add_chart(g_temp, "J11")

g_cidade = BarChart()
g_cidade.type = "col"
g_cidade.title = "Leads por cidade"
g_cidade.height, g_cidade.width = 9, 12
g_cidade.add_data(Reference(dados, min_col=5, min_row=1, max_row=n_cidade + 1), titles_from_data=True)
g_cidade.set_categories(Reference(dados, min_col=4, min_row=2, max_row=n_cidade + 1))
g_cidade.legend = None
g_cidade.dLbls = DataLabelList()
g_cidade.dLbls.showVal = True
dash.add_chart(g_cidade, "B34")

g_builder = BarChart()
g_builder.type = "col"
g_builder.title = "Com que ferramenta o site foi feito"
g_builder.height, g_builder.width = 9, 12
g_builder.add_data(Reference(dados, min_col=11, min_row=1, max_row=n_builder + 1), titles_from_data=True)
g_builder.set_categories(Reference(dados, min_col=10, min_row=2, max_row=n_builder + 1))
g_builder.legend = None
g_builder.dLbls = DataLabelList()
g_builder.dLbls.showVal = True
dash.add_chart(g_builder, "J34")

# ── cenarios de receita (precos reais do painel) ────────────────────────────
dash["B53"] = "Possibilidades de receita"
dash["B53"].font = F_SECAO
dash["B54"] = (
    f"Preços vivos do painel: à vista ${PRECO_FULL:,.0f} · 6x $299 (${PRECO_M6_TOTAL:,.0f}) · "
    f"12x $179 (${PRECO_M12_TOTAL:,.0f}) · suporte ${CARE_MENSAL:,.0f}/mês após {CARE_MESES_INCLUSOS} meses"
)
dash["B54"].font = F_SUB

cab_cen = ["Cenário", "Base de leads", "Conversão", "Fechamentos", "Receita à vista", "Se todos 12x", "Recorrente/mês (após 12m)"]
for i, t in enumerate(cab_cen):
    c = dash.cell(row=56, column=2 + i, value=t)
    c.font = F_CAB
    c.fill = FILL_ESCURO
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDA_FINA
dash.row_dimensions[56].height = 30

cenarios = [
    ("Conservador", "Só os quentes (8-10)", 0.10, len(quentes)),
    ("Provável", "Só os quentes (8-10)", 0.20, len(quentes)),
    ("Bom", "Quentes + mornos (6-10)", 0.15, len(quentes) + len(mornos)),
    ("Otimista", "Quentes + mornos (6-10)", 0.25, len(quentes) + len(mornos)),
]

linha = 57
for nome, base, taxa, universo in cenarios:
    fechamentos = max(1, round(universo * taxa))
    valores = [
        nome,
        f"{universo} leads · {base}",
        taxa,
        fechamentos,
        fechamentos * PRECO_FULL,
        fechamentos * PRECO_M12_TOTAL,
        fechamentos * CARE_MENSAL,
    ]
    for i, v in enumerate(valores):
        c = dash.cell(row=linha, column=2 + i, value=v)
        c.font = F_NEGRITO if i == 0 else F_CORPO
        c.border = BORDA_FINA
        if i == 2:
            c.number_format = "0%"
            c.alignment = Alignment(horizontal="center")
        elif i == 3:
            c.number_format = "#,##0"
            c.alignment = Alignment(horizontal="center")
        elif i >= 4:
            c.number_format = '"$"#,##0'
        if linha % 2 == 1:
            c.fill = FILL_PAR
    linha += 1

dash.cell(row=linha + 1, column=2, value=(
    "Conversão de DM fria em nicho local costuma ficar entre 5% e 15%; os 20-25% assumem prévia pronta "
    "e follow-up. São projeções, não promessas."
)).font = F_SUB

larguras_dash = {"A": 3, "B": 22, "C": 24, "D": 13, "E": 13, "F": 16, "G": 16, "H": 22, "I": 4, "J": 14}
for col_letra, larg in larguras_dash.items():
    dash.column_dimensions[col_letra].width = larg


# ════════════════════════════════════════════════════════════════════════════
# helper de tabela
# ════════════════════════════════════════════════════════════════════════════
def montar_tabela(ws, cabecalhos, linhas_dados, larguras, titulo, subtitulo):
    ws.sheet_view.showGridLines = False
    ws["A1"] = titulo
    ws["A1"].font = F_TITULO
    ws["A2"] = subtitulo
    ws["A2"].font = F_SUB
    ws.row_dimensions[1].height = 28

    for i, t in enumerate(cabecalhos):
        c = ws.cell(row=4, column=1 + i, value=t)
        c.font = F_CAB
        c.fill = FILL_ESCURO
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDA_FINA
    ws.row_dimensions[4].height = 30

    for r, linha_dados in enumerate(linhas_dados, start=5):
        for i, v in enumerate(linha_dados):
            c = ws.cell(row=r, column=1 + i, value=v)
            c.font = F_CORPO
            c.border = BORDA_FINA
            c.alignment = Alignment(vertical="top", wrap_text=(i == len(linha_dados) - 1))
            if r % 2 == 1:
                c.fill = FILL_PAR

    for i, larg in enumerate(larguras):
        ws.column_dimensions[get_column_letter(1 + i)].width = larg

    ultima = 4 + len(linhas_dados)
    ws.auto_filter.ref = f"A4:{get_column_letter(len(cabecalhos))}{ultima}"
    ws.freeze_panes = "A5"
    return ultima


# ════════════════════════════════════════════════════════════════════════════
# LEADS
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Leads")
cab = ["Pot.", "Temp.", "Negócio", "Nicho", "Cidade", "Site (clique)", "Print do site",
       "Nota site", "Feito em", "Booking", "Copyright", "Instagram (clique)", "Seguidores",
       "E-mail (clique)", "Ângulo de abordagem"]
linhas_leads = [
    [l["potencial"], l["temp"], l["negocio"], l["nicho"], l["cidade"], l["site"], "abrir print",
     l["nota_site"], l["builder"].replace("custom/desconhecido", "sob medida"), l["booking"],
     l["ano_copyright"], l["instagram"], l["seg_num"] or "", l["emails"], l["observacoes_angulo"]]
    for l in leads
]
ultima = montar_tabela(
    ws, cab, linhas_leads,
    [6, 11, 30, 15, 16, 30, 13, 9, 14, 9, 11, 24, 11, 34, 62],
    "Leads · carteira completa",
    "Ordenado por potencial. Site, print, Instagram e e-mail são clicáveis: confira cada avaliação você mesmo.",
)

for r, l in zip(range(5, ultima + 1), leads):
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=8).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=13).number_format = "#,##0"
    aplicar_link(ws, r, 6, f"https://{l['site']}")
    if l["screenshot"]:
        aplicar_link(ws, r, 7, l["screenshot"].replace("prospeccao/", ""))
    else:
        ws.cell(row=r, column=7).value = ""
    if l["instagram"]:
        aplicar_link(ws, r, 12, f"https://instagram.com/{l['instagram'].lstrip('@')}")
    if l["emails"]:
        aplicar_link(ws, r, 14, f"mailto:{l['emails'].split(' / ')[0]}")

faixa_pot = f"A5:A{ultima}"
ws.conditional_formatting.add(faixa_pot, CellIsRule(operator="greaterThanOrEqual", formula=["8"], fill=FILL_QUENTE, font=Font(bold=True, color="0A6B3D")))
ws.conditional_formatting.add(faixa_pot, CellIsRule(operator="between", formula=["6", "7"], fill=FILL_MORNO, font=Font(bold=True, color="8A5A12")))
ws.conditional_formatting.add(faixa_pot, CellIsRule(operator="lessThanOrEqual", formula=["5"], fill=FILL_FRIO, font=Font(color=CINZA)))
ws.conditional_formatting.add(f"H5:H{ultima}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=10, color=CIANO))
ws.conditional_formatting.add(f"M5:M{ultima}", DataBarRule(start_type="min", end_type="max", color="B39DDB"))


# ════════════════════════════════════════════════════════════════════════════
# PRIORIDADE (aba de trabalho)
# ════════════════════════════════════════════════════════════════════════════
prio = wb.create_sheet("Prioridade")
alvos = [l for l in leads if l["potencial"] >= 6]
cab_p = ["Pot.", "Negócio", "Nicho", "Cidade", "Site (clique)", "Print do site", "Instagram (clique)",
         "Seguidores", "E-mail (clique)", "Ângulo de abordagem", "Status", "Seguiu em", "Comentou em",
         "DM enviada em", "Notas"]
linhas_p = [
    [l["potencial"], l["negocio"], l["nicho"], l["cidade"], l["site"], "abrir print", l["instagram"],
     l["seg_num"] or "", l["emails"], l["observacoes_angulo"], "A fazer", "", "", "", ""]
    for l in alvos
]
ultima_p = montar_tabela(
    prio, cab_p, linhas_p,
    [6, 30, 15, 16, 28, 13, 24, 11, 32, 58, 14, 12, 13, 15, 40],
    "Prioridade · seus alvos de verdade",
    "Potencial 6+. Fluxo: seguir → comentar 2 ou 3 dias → mandar a DM com a prévia pronta.",
)

status_dv = DataValidation(
    type="list",
    formula1='"A fazer,Seguindo,Comentou,DM enviada,Respondeu,Prévia enviada,Negociando,Ganho,Perdido"',
    allow_blank=True,
)
prio.add_data_validation(status_dv)
status_dv.add(f"K5:K{ultima_p}")

for r, l in zip(range(5, ultima_p + 1), alvos):
    prio.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    prio.cell(row=r, column=8).number_format = "#,##0"
    for c in (12, 13, 14):
        prio.cell(row=r, column=c).number_format = "DD/MM/YYYY"
    aplicar_link(prio, r, 5, f"https://{l['site']}")
    if l["screenshot"]:
        aplicar_link(prio, r, 6, l["screenshot"].replace("prospeccao/", ""))
    else:
        prio.cell(row=r, column=6).value = ""
    if l["instagram"]:
        aplicar_link(prio, r, 7, f"https://instagram.com/{l['instagram'].lstrip('@')}")
    if l["emails"]:
        aplicar_link(prio, r, 9, f"mailto:{l['emails'].split(' / ')[0]}")

prio.conditional_formatting.add(f"A5:A{ultima_p}", CellIsRule(operator="greaterThanOrEqual", formula=["8"], fill=FILL_QUENTE, font=Font(bold=True, color="0A6B3D")))
prio.conditional_formatting.add(f"A5:A{ultima_p}", CellIsRule(operator="between", formula=["6", "7"], fill=FILL_MORNO, font=Font(bold=True, color="8A5A12")))
prio.conditional_formatting.add(f"K5:K{ultima_p}", CellIsRule(operator="equal", formula=['"Ganho"'], fill=PatternFill("solid", fgColor="C8F0D8")))
prio.conditional_formatting.add(f"K5:K{ultima_p}", CellIsRule(operator="equal", formula=['"Perdido"'], fill=PatternFill("solid", fgColor="F5D5D5")))
prio.conditional_formatting.add(f"K5:K{ultima_p}", CellIsRule(operator="equal", formula=['"DM enviada"'], fill=PatternFill("solid", fgColor="DCE9F7")))


# ════════════════════════════════════════════════════════════════════════════
# INSTAGRAM
# ════════════════════════════════════════════════════════════════════════════
ig = wb.create_sheet("Instagram")
contas = sorted([l for l in leads if l["instagram"]], key=lambda l: (-l["potencial"], -l["seg_num"]))
cab_ig = ["Conta", "Seguidores", "Nicho", "Cidade", "Pot. de venda", "Site", "Seguiu?", "Comentou?", "Notas"]
linhas_ig = [
    [l["instagram"], l["seg_num"] or "", l["nicho"], l["cidade"], l["potencial"], l["site"], "", "", ""]
    for l in contas
]
ultima_ig = montar_tabela(
    ig, cab_ig, linhas_ig,
    [28, 12, 16, 16, 13, 30, 10, 12, 40],
    "Instagram · lista de aquecimento",
    "Contas verificadas pelo link no próprio site do negócio. Máx. ~150 ações/dia em conta nova, ou o Instagram trava.",
)

sim_nao = DataValidation(type="list", formula1='"sim,não"', allow_blank=True)
ig.add_data_validation(sim_nao)
sim_nao.add(f"G5:H{ultima_ig}")

for r, l in zip(range(5, ultima_ig + 1), contas):
    ig.cell(row=r, column=2).number_format = "#,##0"
    ig.cell(row=r, column=5).alignment = Alignment(horizontal="center")
    aplicar_link(ig, r, 1, f"https://instagram.com/{l['instagram'].lstrip('@')}")
    aplicar_link(ig, r, 6, f"https://{l['site']}")
ig.conditional_formatting.add(f"B5:B{ultima_ig}", DataBarRule(start_type="min", end_type="max", color=CIANO))
ig.conditional_formatting.add(f"E5:E{ultima_ig}", CellIsRule(operator="greaterThanOrEqual", formula=["8"], fill=FILL_QUENTE, font=Font(bold=True, color="0A6B3D")))


# ════════════════════════════════════════════════════════════════════════════
# AVISOS
# ════════════════════════════════════════════════════════════════════════════
av = wb.create_sheet("Avisos")
av.sheet_view.showGridLines = False
av["A1"] = "Avisos de verificação"
av["A1"].font = F_TITULO
av["A2"] = "Confira estes antes de gastar tempo (ou credibilidade) com eles."
av["A2"].font = F_SUB

avisos = [
    ("Championship Martial Arts", "championshipmartialarts.com",
     "Os e-mails capturados são de OUTRAS franquias (d-a-m.ca, pakskarate.com, randori-pro.com). "
     "Achar o contato da unidade de Orlando antes de abordar."),
    ("Kansas City Tree Services", "kansascitytreeservices.net",
     "Telefone aparente 816-555-7890, padrão clássico de placeholder. Pode ser site de lead-broker "
     "e não negócio real. Confirmar no Google Maps antes."),
    ("Purrfect Grooming", "purrfectgrooming.pet",
     "A sede parece ser Doral/Miami, não Orlando. O site é ótimo alvo, só ajuste a cidade na conversa."),
    ("Redes grandes (nota 1-3)", "autobell.com, tidalwaveautospa.com, superstarcarwashes...",
     "Site bom e agência contratada: não são venda. Servem para seguir e comentar, "
     "porque colocam você no grafo social do nicho."),
    ("E-mails genéricos", "info@, contact@, support@",
     "Chegam na caixa geral e costumam morrer lá. O Instagram continua sendo o canal principal: "
     "e-mail é reforço, não substituto."),
]
for i, t in enumerate(["Negócio", "Site", "O que verificar"]):
    c = av.cell(row=4, column=1 + i, value=t)
    c.font = F_CAB
    c.fill = FILL_ESCURO
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDA_FINA

for r, (nome, site, texto) in enumerate(avisos, start=5):
    for i, v in enumerate((nome, site, texto)):
        c = av.cell(row=r, column=1 + i, value=v)
        c.font = F_NEGRITO if i == 0 else F_CORPO
        c.border = BORDA_FINA
        c.alignment = Alignment(vertical="top", wrap_text=(i == 2))
        if r % 2 == 1:
            c.fill = FILL_PAR
    av.row_dimensions[r].height = 46

for letra, larg in (("A", 28), ("B", 42), ("C", 95)):
    av.column_dimensions[letra].width = larg

wb.save("prospeccao/prospeccao-rvland.xlsx")
print(f"prospeccao-rvland.xlsx: {len(leads)} leads, {len(alvos)} prioritários, {len(contas)} contas de IG")
print(f"KPIs -> quentes {len(quentes)} | mornos {len(mornos)} | com email {len(com_email)} | alcance {alcance:,}")
